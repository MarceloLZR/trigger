"""
infrastructure.excel_exporter
--------------------------------
Exporta un DataFrame a .xlsx con formato mínimo profesional:
encabezados en negrita, ancho de columna autoajustado y filtro automático.
"""
from __future__ import annotations
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.interfaces import IExcelExporter


class ExcelExporter(IExcelExporter):
    def export(self, df: pd.DataFrame, destination_path: str, sheet_name: str = "Resultado") -> str:
        with pd.ExcelWriter(destination_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font

                max_len = max(
                    [len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].head(500)]
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

        return destination_path
